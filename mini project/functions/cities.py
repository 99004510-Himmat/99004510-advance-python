"""
Function to set the cities from the entered PS number from source file
"""
import openpyxl


def cities(row1):
    if 1 <= row1 <= 16:
        object_to_be_loaded = openpyxl.load_workbook("D:\\H\\Work\\Genesis\\99004510-advance python\\mini project\\resource\\raw_data.xlsx")
        sheet_to_be_loaded = object_to_be_loaded['Cities']

        datasheet = []

        workbook = openpyxl.Workbook()
        workbook['Sheet'].title = 'Data'
        sheet_to_be_written = workbook.active
        for i in range(row1, row1 + 1):
            for j in range(1, 22):
                datasheet.append(sheet_to_be_loaded.cell(i, j).value)

        sheet_to_be_written.append(datasheet)
        workbook.save('results.xlsx')
        print('Data saved')

        return True
    else:
        return False
